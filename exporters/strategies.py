"""
Export strategies using Strategy pattern.
"""
from abc import ABC, abstractmethod
from typing import List, Dict, Any
import csv
import xml.etree.ElementTree as ET
import xml.dom.minidom
from datetime import datetime
from tabulate import tabulate
from models import Game

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportError(Exception):
    """Export error exception"""


class ExportStrategy(ABC):
    """Abstract base class for export strategies"""

    @abstractmethod
    def export(self, games: List[Game], filepath: str) -> None:
        """Export games to file"""
        pass

    @property
    @abstractmethod
    def extension(self) -> str:
        """File extension for this export format"""
        pass

    @property
    @abstractmethod
    def is_available(self) -> bool:
        """Check if export format is available"""
        pass


class CSVExportStrategy(ExportStrategy):
    """CSV export strategy"""

    @property
    def extension(self) -> str:
        return ".csv"

    @property
    def is_available(self) -> bool:
        return True

    def export(self, games: List[Game], filepath: str) -> None:
        """Export games to CSV file"""
        if not games:
            raise ExportError("No games to export")

        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            # Determine fields based on first game
            sample_dict = games[0].to_dict()
            fieldnames = list(sample_dict.keys())

            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for game in games:
                writer.writerow(game.to_dict())


class MarkdownExportStrategy(ExportStrategy):
    """Markdown export strategy"""

    @property
    def extension(self) -> str:
        return ".md"

    @property
    def is_available(self) -> bool:
        return True

    def export(self, games: List[Game], filepath: str) -> None:
        """Export games to Markdown file"""
        if not games:
            raise ExportError("No games to export")

        with open(filepath, 'w', encoding='utf-8') as mdfile:
            # Header
            mdfile.write("# Games Database Results\n\n")
            mdfile.write(f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            mdfile.write(f"Total games found: {len(games)}\n\n")

            # Prepare table data
            sample_dict = games[0].to_dict()
            headers = list(sample_dict.keys())
            table_data = [list(game.to_dict().values()) for game in games]

            # Generate markdown table
            markdown_table = tabulate(table_data, headers=headers, tablefmt="github")
            mdfile.write(markdown_table)


class XLSXExportStrategy(ExportStrategy):
    """XLSX export strategy"""

    @property
    def extension(self) -> str:
        return ".xlsx"

    @property
    def is_available(self) -> bool:
        return OPENPYXL_AVAILABLE

    def export(self, games: List[Game], filepath: str) -> None:
        """Export games to XLSX file"""
        if not OPENPYXL_AVAILABLE:
            raise ExportError("openpyxl non è installato")

        if not games:
            raise ExportError("No games to export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Games Results"

        # Get headers from first game
        sample_dict = games[0].to_dict()
        headers = list(sample_dict.keys())

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for row, game in enumerate(games, 2):
            game_dict = game.to_dict()
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=game_dict[header])

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = ws.column_dimensions[chr(64 + col)]
            for row in range(1, len(games) + 2):
                try:
                    cell_value = str(ws.cell(row=row, column=col).value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            column.width = adjusted_width

        wb.save(filepath)


class XMLExportStrategy(ExportStrategy):
    """XML export strategy"""

    @property
    def extension(self) -> str:
        return ".xml"

    @property
    def is_available(self) -> bool:
        return True

    def export(self, games: List[Game], filepath: str) -> None:
        """Export games to XML file"""
        if not games:
            raise ExportError("No games to export")

        root = ET.Element("games_database")

        # Metadata
        metadata = ET.SubElement(root, "metadata")
        ET.SubElement(metadata, "export_date").text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ET.SubElement(metadata, "total_games").text = str(len(games))

        # Check if AI reviews are present
        has_ai_reviews = any(game.ai_review is not None for game in games)
        ET.SubElement(metadata, "includes_ai_reviews").text = str(has_ai_reviews).lower()

        # Games
        games_element = ET.SubElement(root, "games")

        for game in games:
            game_element = ET.SubElement(games_element, "game")
            ET.SubElement(game_element, "title").text = game.title
            ET.SubElement(game_element, "platforms").text = game.platforms_str
            ET.SubElement(game_element, "release_date").text = game.release_date or "N/A"
            ET.SubElement(game_element, "genres").text = game.genres_str

            if has_ai_reviews:
                ET.SubElement(game_element, "ai_review").text = game.ai_review or "N/A"
                ET.SubElement(game_element, "ai_rating").text = game.ai_rating_str

        # Pretty print XML
        rough_string = ET.tostring(root, encoding='unicode')
        reparsed = xml.dom.minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ")

        # Remove extra empty lines
        lines = [line for line in pretty_xml.split('\n') if line.strip()]

        with open(filepath, 'w', encoding='utf-8') as xmlfile:
            xmlfile.write('\n'.join(lines))


class ExportStrategyFactory:
    """Factory for creating export strategies"""

    _strategies = {
        "CSV": CSVExportStrategy,
        "Markdown": MarkdownExportStrategy,
        "XLSX (Excel)": XLSXExportStrategy,
        "XML": XMLExportStrategy
    }

    @classmethod
    def get_available_formats(cls) -> List[str]:
        """Get list of available export formats"""
        formats = []
        for name, strategy_class in cls._strategies.items():
            strategy = strategy_class()
            if strategy.is_available:
                formats.append(name)
            else:
                formats.append(f"{name} - Non disponibile")
        return formats

    @classmethod
    def create_strategy(cls, format_name: str) -> ExportStrategy:
        """Create export strategy by format name"""
        # Remove " - Non disponibile" suffix if present
        clean_name = format_name.split(" -")[0]

        if clean_name not in cls._strategies:
            raise ExportError(f"Unknown export format: {format_name}")

        strategy = cls._strategies[clean_name]()

        if not strategy.is_available:
            raise ExportError(f"Export format {clean_name} is not available")

        return strategy
