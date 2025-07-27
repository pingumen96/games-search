





import requests
from tabulate import tabulate
import questionary
from datetime import datetime
import os
import csv
import traceback
from dotenv import load_dotenv
import xml.etree.ElementTree as ET
import xml.dom.minidom
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class GamesDB:
    def __init__(self):
        load_dotenv()
        self.api_key = os.getenv("RAWG_API_KEY")
        if not self.api_key:
            raise ValueError("RAWG_API_KEY non trovata nel file .env")
        self.games_data = []
        self.current_games = []

    def get_year_input(self):
        """Chiede l'anno di ricerca con validazione"""
        current_year = datetime.now().year

        while True:
            year_input = questionary.text(
                f"Inserisci anno (1970-{current_year}) o premi Invio per anno corrente ({current_year}):"
            ).ask()

            if not year_input:  # Se premi solo Invio
                return current_year

            try:
                year = int(year_input)
                if 1970 <= year <= current_year:
                    return year
                else:
                    print(f"Errore: L'anno deve essere tra 1970 e {current_year}")
            except ValueError:
                print("Errore: Inserisci un numero valido")

    def get_month_input(self):
        """Chiede il mese di ricerca con validazione"""
        current_month = datetime.now().month

        while True:
            month_input = questionary.text(
                f"Inserisci mese (1-12) o premi Invio per mese corrente ({current_month}):"
            ).ask()

            if not month_input:  # Se premi solo Invio
                return current_month

            try:
                month = int(month_input)
                if 1 <= month <= 12:
                    return month
                else:
                    print("Errore: Il mese deve essere tra 1 e 12")
            except ValueError:
                print("Errore: Inserisci un numero valido")

    def get_platform_filter(self):
        """Chiede il filtro per piattaforma"""
        common_platforms = [
            "PC", "PlayStation", "Xbox", "Nintendo Switch", "iOS", "Android",
            "NES", "SNES", "Game Boy", "Arcade", "Atari", "Sega Genesis"
        ]

        choice = questionary.select(
            "Vuoi filtrare per piattaforme?",
            choices=[
                "No, mostra tutte le piattaforme",
                "Sì, scegli da lista comuni",
                "Sì, inserimento manuale"
            ]
        ).ask()

        if choice == "No, mostra tutte le piattaforme":
            return None
        elif choice == "Sì, scegli da lista comuni":
            selected = questionary.checkbox(
                "Seleziona le piattaforme (spazio per selezionare, Invio per confermare):",
                choices=common_platforms
            ).ask()
            return selected if selected else None
        else:
            manual_input = questionary.text(
                "Inserisci piattaforme separate da virgola (es: NES, SNES, Game Boy):"
            ).ask()
            if manual_input:
                return [p.strip() for p in manual_input.split(',')]
            return None

    def fetch_games(self, year, month, platform_filter=None):
        """Recupera i giochi dall'API RAWG"""
        # Build date range
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year+1}-01-01"
        else:
            end_date = f"{year}-{month+1:02d}-01"

        print(f"\n🔍 Cercando giochi rilasciati nel {month}/{year}...")

        # Call RAWG API
        url = f"https://api.rawg.io/api/games?dates={start_date},{end_date}&page_size=100&key={self.api_key}"

        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
        except requests.exceptions.RequestException as e:
            print(f"Errore nella richiesta API: {e}")
            return []

        # Extract games data
        games = []
        for game in data.get('results', []):
            title = game.get('name')
            release_date = game.get('released')
            platforms_data = game.get('platforms')
            if not isinstance(platforms_data, list):
                platforms_data = []
            platforms = [p['platform']['name'] for p in platforms_data]
            genres = [g['name'] for g in game.get('genres', [])]

            # Apply platform filter if specified
            if platform_filter:
                if not any(filter_platform.lower() in platform.lower()
                          for platform in platforms
                          for filter_platform in platform_filter):
                    continue

            games.append({
                "Title": title,
                "Platforms": ", ".join(platforms),
                "Release Date": release_date,
                "Genres": ", ".join(genres)
            })

        self.games_data = games
        return games

    def sort_games(self, games):
        """Ordina i risultati secondo la scelta dell'utente"""
        if not games:
            return games

        sort_choice = questionary.select(
            "Come vuoi ordinare i risultati?",
            choices=[
                "Data di rilascio",
                "Titolo (A-Z)",
                "Titolo (Z-A)",
                "Piattaforma",
                "Non ordinare"
            ]
        ).ask()

        if sort_choice == "Data di rilascio":
            return sorted(games, key=lambda x: x['Release Date'] or '9999-12-31')
        elif sort_choice == "Titolo (A-Z)":
            return sorted(games, key=lambda x: x['Title'].lower())
        elif sort_choice == "Titolo (Z-A)":
            return sorted(games, key=lambda x: x['Title'].lower(), reverse=True)
        elif sort_choice == "Piattaforma":
            return sorted(games, key=lambda x: x['Platforms'])
        else:
            return games

    def display_results(self, games):
        """Mostra i risultati in tabella"""
        if not games:
            print("\n❌ Nessun gioco trovato per i criteri specificati.")
            return None

        sorted_games = self.sort_games(games)
        self.current_games = sorted_games

        # Prepara i dati per tabulate
        headers = ["Titolo", "Piattaforme", "Data Rilascio", "Generi"]
        table_data = []

        for game in sorted_games:
            table_data.append([
                game["Title"],
                game["Platforms"],
                game["Release Date"] or "N/A",
                game["Genres"]
            ])

        print(f"\n📋 Trovati {len(sorted_games)} giochi:")
        print("=" * 120)
        print(tabulate(table_data, headers=headers, tablefmt="pretty"))

        return sorted_games

    def export_data(self, games):
        """Esporta i risultati in vari formati"""
        if not games:
            print("❌ Nessun dato da esportare!")
            return

        export = questionary.confirm("Vuoi esportare i risultati?").ask()

        if not export:
            return

        # Scelta del formato
        format_choice = questionary.select(
            "Scegli il formato di esportazione:",
            choices=[
                "CSV",
                "Markdown",
                "XLSX (Excel)" if OPENPYXL_AVAILABLE else "XLSX (Excel) - Non disponibile",
                "XML"
            ]
        ).ask()

        # Se XLSX non è disponibile
        if format_choice.startswith("XLSX") and not OPENPYXL_AVAILABLE:
            print("❌ La libreria openpyxl non è installata. Installa con: pip install openpyxl")
            return

        # Rimuovi la parte " - Non disponibile" se presente
        format_choice = format_choice.split(" -")[0]

        filename = questionary.text(
            "Nome del file (senza estensione):",
            default="games_results"
        ).ask()

        if not filename:
            filename = "games_results"

        # Determina l'estensione del file
        extensions = {
            "CSV": ".csv",
            "Markdown": ".md",
            "XLSX (Excel)": ".xlsx",
            "XML": ".xml"
        }
        
        filepath = f"{filename}{extensions[format_choice]}"

        try:
            if format_choice == "CSV":
                self._export_csv(games, filepath)
            elif format_choice == "Markdown":
                self._export_markdown(games, filepath)
            elif format_choice == "XLSX (Excel)":
                self._export_xlsx(games, filepath)
            elif format_choice == "XML":
                self._export_xml(games, filepath)

            print(f"✅ File esportato con successo: {os.path.abspath(filepath)}")
        except Exception as e:
            print(f"❌ Errore durante l'esportazione: {e}")

    def _export_csv(self, games, filepath):
        """Esporta in formato CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ["Title", "Platforms", "Release Date", "Genres"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for game in games:
                writer.writerow(game)

    def _export_markdown(self, games, filepath):
        """Esporta in formato Markdown"""
        with open(filepath, 'w', encoding='utf-8') as mdfile:
            # Intestazione
            mdfile.write("# Games Database Results\n\n")
            mdfile.write(f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            mdfile.write(f"Total games found: {len(games)}\n\n")
            
            # Tabella
            headers = ["Title", "Platforms", "Release Date", "Genres"]
            table_data = []
            
            for game in games:
                table_data.append([
                    game["Title"],
                    game["Platforms"],
                    game["Release Date"] or "N/A",
                    game["Genres"]
                ])
            
            # Usa tabulate per generare la tabella markdown
            markdown_table = tabulate(table_data, headers=headers, tablefmt="github")
            mdfile.write(markdown_table)

    def _export_xlsx(self, games, filepath):
        """Esporta in formato XLSX (Excel)"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl non è installato")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Games Results"
        
        # Headers
        headers = ["Title", "Platforms", "Release Date", "Genres"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, game in enumerate(games, 2):
            ws.cell(row=row, column=1, value=game["Title"])
            ws.cell(row=row, column=2, value=game["Platforms"])
            ws.cell(row=row, column=3, value=game["Release Date"] or "N/A")
            ws.cell(row=row, column=4, value=game["Genres"])
        
        # Auto-adjust column widths
        for col in range(1, 5):
            max_length = 0
            column = ws.column_dimensions[chr(64 + col)]
            for row in range(1, len(games) + 2):
                try:
                    cell_value = str(ws.cell(row=row, column=col).value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            column.width = adjusted_width
        
        wb.save(filepath)

    def _export_xml(self, games, filepath):
        """Esporta in formato XML"""
        root = ET.Element("games_database")
        
        # Metadata
        metadata = ET.SubElement(root, "metadata")
        ET.SubElement(metadata, "export_date").text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ET.SubElement(metadata, "total_games").text = str(len(games))
        
        # Games
        games_element = ET.SubElement(root, "games")
        
        for game in games:
            game_element = ET.SubElement(games_element, "game")
            ET.SubElement(game_element, "title").text = game["Title"]
            ET.SubElement(game_element, "platforms").text = game["Platforms"]
            ET.SubElement(game_element, "release_date").text = game["Release Date"] or "N/A"
            ET.SubElement(game_element, "genres").text = game["Genres"]
        
        # Pretty print XML
        rough_string = ET.tostring(root, encoding='unicode')
        reparsed = xml.dom.minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ")
        
        # Rimuovi righe vuote extra
        lines = [line for line in pretty_xml.split('\n') if line.strip()]
        
        with open(filepath, 'w', encoding='utf-8') as xmlfile:
            xmlfile.write('\n'.join(lines))

    def search_games(self):
        """Funzione principale per la ricerca giochi"""
        print("\n🎮 === RICERCA GIOCHI ===")

        # Get search parameters
        year = self.get_year_input()
        month = self.get_month_input()
        platform_filter = self.get_platform_filter()

        # Fetch and display results
        games = self.fetch_games(year, month, platform_filter)
        result_games = self.display_results(games)

        # Export option
        if result_games:
            self.export_data(result_games)

    def show_main_menu(self):
        """Mostra il menu principale"""
        while True:
            print("\n" + "="*50)
            print("🎮 GAMES DATABASE - Menu Principale")
            print("="*50)

            choice = questionary.select(
                "Scegli un'azione:",
                choices=[
                    "🔍 Cerca giochi per data",
                    "📊 Mostra ultimi risultati",
                    "📁 Esporta ultimi risultati",
                    "❌ Esci"
                ]
            ).ask()

            if choice == "🔍 Cerca giochi per data":
                self.search_games()
            elif choice == "📊 Mostra ultimi risultati":
                if self.current_games:
                    self.display_results(self.current_games)
                else:
                    print("\n❌ Nessun risultato precedente disponibile. Effettua prima una ricerca!")
            elif choice == "📁 Esporta ultimi risultati":
                self.export_data(self.current_games)
            elif choice == "❌ Esci":
                print("\n👋 Arrivederci!")
                break

def main():
    """Funzione principale"""
    try:
        app = GamesDB()
        app.show_main_menu()
    except KeyboardInterrupt:
        print("\n\n👋 Uscita forzata. Arrivederci!")
    except Exception as e:
        print(f"\n❌ Errore imprevisto: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
